#!/usr/bin/env python3
"""
Faliz 3D Web Development Progress Tracker
================================================================================
PROJECT: Faliz 3D Web Dev Progress Tracker
DESCRIPTION: Generates a professionally styled Excel (.xlsx) spreadsheet to track
             learner progress through Faliz's 3-stage roadmap for building
             interactive 3D websites with Three.js, React Three Fiber, GSAP,
             and physics engines.
PRIMARY GOAL: Empower aspiring 3D creative developers with a structured,
              motivational tracking tool that follows Faliz's step-by-step
              tutoring philosophy. Includes performance optimization reminders.
================================================================================

Tech Stack: Python 3.11 + openpyxl
Run: python faliz_3d_progress_tracker.py
Output: faliz_3d_progress_tracker.xlsx (ready to use in Excel / Google Sheets)
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment

def create_faliz_tracker():
    """Main function to build the complete Faliz 3D Progress Tracker workbook."""
    try:
        wb = Workbook()

        # ============================================
        # SHEET 1: Progress Tracker (Main Roadmap)
        # ============================================
        ws = wb.active
        ws.title = "Progress Tracker"

        # Define styles
        title_font = Font(name='Arial', size=22, bold=True, color='00FFFF')
        title_fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
        phase_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
        phase_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        alt_row_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Title Section
        ws.merge_cells('A1:G1')
        ws['A1'] = "🚀 FALIZ 3D WEB DEVELOPMENT PROGRESS TRACKER"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Subtitle / Motivation
        ws.merge_cells('A2:G2')
        ws['A2'] = "Master interactive 3D websites step-by-step • Track your journey from first cube to full physics-powered experiences • Performance-first mindset from Day 1"
        ws['A2'].font = Font(name='Arial', size=10, italic=True, color='1E3A5F')
        ws['A2'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[2].height = 28

        # Faliz Quote
        ws.merge_cells('A3:G3')
        ws['A3'] = '"Every legendary 3D website begins with a single BoxGeometry. Stay curious, optimize early, and build what excites you." — Faliz'
        ws['A3'].font = Font(name='Arial', size=9, italic=True, color='00BFFF')
        ws['A3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[3].height = 22

        # Headers
        headers = ["Phase", "Lesson / Topic", "Description & Key Skills", "Status", "Progress %", "Notes / Reflections", "Resources & Links"]
        header_row = 5
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 25

        # Curriculum Data - Complete, no placeholders
        curriculum = [
            {
                "phase": "PHASE 1: THE SCENE & CAMERA",
                "lessons": [
                    {
                        "topic": "1.1 Scene, Renderer & Perspective Camera Setup",
                        "desc": "Create Scene(), WebGLRenderer(), PerspectiveCamera(). Handle window resize, set pixelRatio for mobile. Understand coordinate system.",
                        "resources": "threejs.org/docs/#api/en/scenes/Scene | threejs.org/docs/#api/en/cameras/PerspectiveCamera"
                    },
                    {
                        "topic": "1.2 Adding Basic Meshes & Geometries",
                        "desc": "BoxGeometry, SphereGeometry, PlaneGeometry + Mesh + MeshPhongMaterial. Experiment with position, rotation, scale.",
                        "resources": "threejs.org/docs/#api/en/geometries/BoxGeometry"
                    },
                    {
                        "topic": "1.3 Camera Controls & OrbitControls",
                        "desc": "Add interactive camera with mouse drag/zoom. Vanilla OrbitControls or @react-three/drei <OrbitControls />.",
                        "resources": "threejs.org/examples/#misc_controls_orbit"
                    },
                    {
                        "topic": "1.4 The Animation Loop (requestAnimationFrame)",
                        "desc": "Build the render loop. Update animations, rotate meshes, render() call. Monitor with stats.js or built-in.",
                        "resources": "developer.mozilla.org/en-US/docs/Web/API/window/requestAnimationFrame"
                    },
                    {
                        "topic": "1.5 Performance Foundations & Mobile Optimization",
                        "desc": "Keep draw calls low, reuse geometries/materials, limit polygon count (<50k for mobile). Test on low-end devices.",
                        "resources": "threejs.org/docs/#manual/en/introduction/Performance"
                    },
                ]
            },
            {
                "phase": "PHASE 2: MATERIALS & LIGHTING",
                "lessons": [
                    {
                        "topic": "2.1 Mesh Materials Deep Dive",
                        "desc": "MeshBasicMaterial vs MeshPhongMaterial vs MeshStandardMaterial. Metalness, roughness, emissive maps. When to use each.",
                        "resources": "threejs.org/docs/#api/en/materials/MeshStandardMaterial"
                    },
                    {
                        "topic": "2.2 Dynamic Lighting (Ambient, Directional, Point, Hemisphere)",
                        "desc": "Add multiple lights. Shadows with DirectionalLight + shadow map. Light helpers for debugging.",
                        "resources": "threejs.org/docs/#api/en/lights/DirectionalLight"
                    },
                    {
                        "topic": "2.3 Textures, UV Mapping & Environment Maps",
                        "desc": "Load textures with TextureLoader. Repeat, offset, anisotropy. Cube textures for reflections.",
                        "resources": "threejs.org/docs/#api/en/loaders/TextureLoader"
                    },
                    {
                        "topic": "2.4 Loading 3D Models (GLTF/GLB) with GLTFLoader",
                        "desc": "Import Blender/Spline exports. Traverse scene, apply materials, optimize meshes. Draco compression.",
                        "resources": "threejs.org/docs/#examples/en/loaders/GLTFLoader | gltf-transform.dev"
                    },
                    {
                        "topic": "2.5 Lighting Baking & Texture Optimization",
                        "desc": "Bake lights in Blender for performance. Compress textures (basis, ktx2). Keep file sizes small for web.",
                        "resources": "docs.blender.org/manual/en/latest/render/eevee/render-settings/indirect-lighting.html"
                    },
                ]
            },
            {
                "phase": "PHASE 3: PHYSICS & INTERACTIVITY",
                "lessons": [
                    {
                        "topic": "3.1 GSAP Scroll-Triggered Camera & Timeline Animations",
                        "desc": "Use GSAP + ScrollTrigger to animate camera position/rotation on scroll. Smooth cinematic storytelling.",
                        "resources": "greensock.com/scrolltrigger/ | threejs + gsap examples"
                    },
                    {
                        "topic": "3.2 Rapier Physics Engine Integration (R3F or Vanilla)",
                        "desc": "Add rigid bodies, colliders, gravity. Mesh physics with @react-three/rapier or cannon-es. Performance tips.",
                        "resources": "rapier.rs/docs/user_guides/javascript | pmndrs.github.io/react-three-rapier"
                    },
                    {
                        "topic": "3.3 Raycasting for Mouse Interaction & Selection",
                        "desc": "THREE.Raycaster + mouse events. Click to select, hover effects, drag objects. Essential for UI/UX in 3D.",
                        "resources": "threejs.org/docs/#api/en/core/Raycaster"
                    },
                    {
                        "topic": "3.4 Post-Processing Effects & Bloom (drei + drei/postprocessing)",
                        "desc": "Add bloom, SSAO, depth of field for cinematic feel. Use EffectComposer or @react-three/postprocessing.",
                        "resources": "pmndrs.github.io/react-three-postprocessing"
                    },
                    {
                        "topic": "3.5 Deployment, Optimization & Production Polish",
                        "desc": "Build for web (Vite + React Three Fiber). Deploy to Vercel/Netlify. Lighthouse audit, mobile testing, final polish.",
                        "resources": "vercel.com | threejs-journey.com"
                    },
                ]
            },
        ]

        current_row = 6
        data_start_row = 6
        all_lessons_count = 0

        for phase_block in curriculum:
            # Phase header row (merged)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            phase_cell = ws.cell(row=current_row, column=1, value=phase_block["phase"])
            phase_cell.font = phase_font
            phase_cell.fill = phase_fill
            phase_cell.alignment = Alignment(horizontal='left', vertical='center')
            phase_cell.border = thin_border
            for col in range(2, 8):
                ws.cell(row=current_row, column=col).border = thin_border
            ws.row_dimensions[current_row].height = 24
            current_row += 1

            # Lessons for this phase
            for lesson in phase_block["lessons"]:
                ws.cell(row=current_row, column=1, value="").border = thin_border  # Phase already shown above
                ws.cell(row=current_row, column=2, value=lesson["topic"]).font = data_font
                ws.cell(row=current_row, column=2).alignment = left_align
                ws.cell(row=current_row, column=2).border = thin_border

                ws.cell(row=current_row, column=3, value=lesson["desc"]).font = data_font
                ws.cell(row=current_row, column=3).alignment = left_align
                ws.cell(row=current_row, column=3).border = thin_border

                status_cell = ws.cell(row=current_row, column=4, value="Not Started")
                status_cell.font = data_font
                status_cell.alignment = center_align
                status_cell.border = thin_border

                progress_cell = ws.cell(row=current_row, column=5, value=0)
                progress_cell.font = data_font
                progress_cell.alignment = center_align
                progress_cell.border = thin_border
                progress_cell.number_format = '0"%"'

                ws.cell(row=current_row, column=6, value="").font = data_font
                ws.cell(row=current_row, column=6).alignment = left_align
                ws.cell(row=current_row, column=6).border = thin_border

                ws.cell(row=current_row, column=7, value=lesson["resources"]).font = Font(name='Arial', size=9, color='0066CC')
                ws.cell(row=current_row, column=7).alignment = left_align
                ws.cell(row=current_row, column=7).border = thin_border

                # Alternating row colors for readability
                if (current_row - data_start_row) % 2 == 1:
                    for col in range(1, 8):
                        if col != 1:  # keep phase header style clean
                            ws.cell(row=current_row, column=col).fill = alt_row_fill

                current_row += 1
                all_lessons_count += 1

        data_end_row = current_row - 1

        # Data Validation for Status column (dropdown)
        status_dv = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Completed"',
            allow_blank=False,
            showDropDown=False
        )
        status_dv.error = "Please choose a valid status from the dropdown list"
        status_dv.errorTitle = "Invalid Status"
        status_dv.prompt = "Select current status"
        status_dv.promptTitle = "Lesson Status"
        ws.add_data_validation(status_dv)
        status_dv.add(f'D7:D{data_end_row}')

        # Conditional Formatting for Status column (color coding)
        # Completed = green
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        green_font = Font(color='006100')
        ws.conditional_formatting.add(
            f'D7:D{data_end_row}',
            FormulaRule(formula=['$D7="Completed"'], fill=green_fill, font=green_font)
        )
        # In Progress = yellow
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        yellow_font = Font(color='9C5700')
        ws.conditional_formatting.add(
            f'D7:D{data_end_row}',
            FormulaRule(formula=['$D7="In Progress"'], fill=yellow_fill, font=yellow_font)
        )
        # Not Started = light red
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006')
        ws.conditional_formatting.add(
            f'D7:D{data_end_row}',
            FormulaRule(formula=['$D7="Not Started"'], fill=red_fill, font=red_font)
        )

        # Column widths
        col_widths = [28, 38, 55, 14, 12, 30, 45]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze panes (freeze header)
        ws.freeze_panes = 'A6'

        # Add helpful comments / tips in a few cells
        comment = Comment("Pro tip from Faliz: Always test your scene on a real mobile device early. Frame rate matters more than visual fidelity on the web.", "Faliz")
        ws['C7'].comment = comment

        # ============================================
        # SHEET 2: Dashboard (Live Progress Overview)
        # ============================================
        ws_dash = wb.create_sheet("Dashboard")

        # Dashboard Title
        ws_dash.merge_cells('A1:F1')
        ws_dash['A1'] = "📊 FALIZ DASHBOARD - YOUR 3D MASTERY OVERVIEW"
        ws_dash['A1'].font = Font(name='Arial', size=18, bold=True, color='00FFFF')
        ws_dash['A1'].fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
        ws_dash['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dash.row_dimensions[1].height = 32

        ws_dash.merge_cells('A2:F2')
        ws_dash['A2'] = "Formulas auto-update when you change Status or Progress % on the Progress Tracker sheet. Keep learning!"
        ws_dash['A2'].font = Font(name='Arial', size=10, italic=True)
        ws_dash['A2'].alignment = Alignment(horizontal='center')

        # Summary Stats Section
        ws_dash['A4'] = "OVERALL PROGRESS"
        ws_dash['A4'].font = Font(name='Arial', size=12, bold=True, color='1E3A5F')

        ws_dash['A5'] = "Total Lessons in Roadmap:"
        ws_dash['B5'] = all_lessons_count
        ws_dash['B5'].font = Font(name='Arial', size=11, bold=True)

        ws_dash['A6'] = "Lessons Completed:"
        ws_dash['B6'] = f'=COUNTIF(\'Progress Tracker\'!D7:D{data_end_row},"Completed")'
        ws_dash['B6'].font = Font(name='Arial', size=11, bold=True, color='006100')

        ws_dash['A7'] = "Lessons In Progress:"
        ws_dash['B7'] = f'=COUNTIF(\'Progress Tracker\'!D7:D{data_end_row},"In Progress")'
        ws_dash['B7'].font = Font(name='Arial', size=11, bold=True, color='9C5700')

        ws_dash['A8'] = "Not Started Yet:"
        ws_dash['B8'] = f'=COUNTIF(\'Progress Tracker\'!D7:D{data_end_row},"Not Started")'
        ws_dash['B8'].font = Font(name='Arial', size=11, bold=True, color='9C0006')

        ws_dash['A10'] = "COMPLETION RATE"
        ws_dash['A10'].font = Font(name='Arial', size=12, bold=True, color='1E3A5F')

        ws_dash['A11'] = "Overall Progress %:"
        ws_dash.merge_cells('B11:C11')
        ws_dash['B11'] = f'=IF(B5>0,ROUND(B6/B5*100,1)&"%","0%")'
        ws_dash['B11'].font = Font(name='Arial', size=16, bold=True, color='00BFFF')
        ws_dash['B11'].alignment = Alignment(horizontal='center')

        # Phase Progress Section
        ws_dash['A13'] = "PHASE BREAKDOWN (Update Status on Tracker sheet to see live changes)"
        ws_dash['A13'].font = Font(name='Arial', size=11, bold=True, color='1E3A5F')
        ws_dash.merge_cells('A13:F13')

        # Phase 1 stats
        ws_dash['A15'] = "Phase 1: Scene & Camera"
        ws_dash['A15'].font = Font(bold=True)
        ws_dash['A16'] = "Completed in Phase 1:"
        ws_dash['B16'] = '=COUNTIF(\'Progress Tracker\'!D7:D11,"Completed")'
        ws_dash['C16'] = "lessons"

        ws_dash['A17'] = "Phase 2: Materials & Lighting"
        ws_dash['A17'].font = Font(bold=True)
        ws_dash['A18'] = "Completed in Phase 2:"
        ws_dash['B18'] = '=COUNTIF(\'Progress Tracker\'!D13:D17,"Completed")'
        ws_dash['C18'] = "lessons"

        ws_dash['A19'] = "Phase 3: Physics & Interactivity"
        ws_dash['A19'].font = Font(bold=True)
        ws_dash['A20'] = "Completed in Phase 3:"
        ws_dash['B20'] = '=COUNTIF(\'Progress Tracker\'!D19:D23,"Completed")'
        ws_dash['C20'] = "lessons"

        # Motivational footer
        ws_dash.merge_cells('A22:F24')
        ws_dash['A22'] = "🔥 NEXT STEPS FROM FALIZ: Pick ONE lesson today. Update its status. Celebrate small wins. The web's most stunning 3D experiences are built by developers who ship consistently. You are next."
        ws_dash['A22'].font = Font(name='Arial', size=10, italic=True, color='1E3A5F')
        ws_dash['A22'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Column widths for dashboard
        ws_dash.column_dimensions['A'].width = 32
        ws_dash.column_dimensions['B'].width = 18
        ws_dash.column_dimensions['C'].width = 12
        ws_dash.column_dimensions['D'].width = 15
        ws_dash.column_dimensions['E'].width = 15
        ws_dash.column_dimensions['F'].width = 15

        # ============================================
        # Final touches
        # ============================================
        # Set print area and page setup for both sheets (nice when printed)
        ws.print_title_rows = '5:5'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True

        ws_dash.page_setup.orientation = 'portrait'

        # Save the workbook
        output_path = "/home/workdir/artifacts/faliz_3d_progress_tracker.xlsx"
        wb.save(output_path)
        print(f"✅ Successfully created: {output_path}")
        print(f"   Total lessons tracked: {all_lessons_count}")
        print("   Sheets: 'Progress Tracker' + 'Dashboard'")
        print("   Ready for you to start your Faliz journey!")

        return output_path

    except Exception as e:
        print(f"❌ Error creating tracker: {str(e)}", file=sys.stderr)
        raise

if __name__ == "__main__":
    create_faliz_tracker()
